import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment

def main():
    file_path = 'Zaloga_Avtomatizirana.xlsx'
    
    # Load workbook for writing (preserves formulas)
    wb = openpyxl.load_workbook(file_path)
    
    # Load workbook for reading (to get evaluated values if needed, though we can just read from wb)
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    
    # Build a lookup dictionary for Naziv from TRENUTNA ZALOGA
    ws_zaloga = wb_data['TRENUTNA ZALOGA']
    naziv_dict = {}
    for row in ws_zaloga.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            naziv_dict[str(row[0])] = row[1]
            
    # Read PRISPELO
    ws_prispelo = wb_data['PRISPELO']
    # Columns: 0:Datum, 1:Šifra, 2:Naziv, 3:Količina, 4:Št. Dobavnice, 5:Dobavitelj, 6:Opomba
    prispelo_data = []
    for row in ws_prispelo.iter_rows(min_row=2, values_only=True):
        datum = row[0]
        if datum is None:
            continue
            
        # Parse datum or check if it's datetime
        is_july = False
        if isinstance(datum, datetime):
            if datum.month == 7 and datum.year == 2026:
                is_july = True
        elif isinstance(datum, str):
            if "2026-07" in datum or "07.2026" in datum or ".7.2026" in datum:
                is_july = True
                
        if is_july:
            sifra = str(row[1]) if row[1] else ""
            naziv = naziv_dict.get(sifra, row[2] or "")
            kolicina = row[3]
            dodatno = f"Dobavnica: {row[4] or ''}, Dobavitelj: {row[5] or ''}"
            opomba = row[6] or ""
            prispelo_data.append([datum, "PRISPELO", sifra, naziv, kolicina, dodatno, opomba])
            
    # Read ODPIS
    ws_odpis = wb_data['ODPIS']
    # Columns: 0:Datum, 1:Šifra, 2:Naziv, 3:Količina, 4:Delovni Nalog, 5:Oseba, 6:Opomba
    odpis_data = []
    for row in ws_odpis.iter_rows(min_row=2, values_only=True):
        datum = row[0]
        if datum is None:
            continue
            
        # Parse datum
        is_july = False
        if isinstance(datum, datetime):
            if datum.month == 7 and datum.year == 2026:
                is_july = True
        elif isinstance(datum, str):
            if "2026-07" in datum or "07.2026" in datum or ".7.2026" in datum:
                is_july = True
                
        if is_july:
            sifra = str(row[1]) if row[1] else ""
            naziv = naziv_dict.get(sifra, row[2] or "")
            kolicina = row[3]
            dodatno = f"DN: {row[4] or ''}, Oseba: {row[5] or ''}"
            opomba = row[6] or ""
            odpis_data.append([datum, "ODPIS", sifra, naziv, kolicina, dodatno, opomba])

    # Combine and sort by date
    all_data = prispelo_data + odpis_data
    # sort by datum (assuming it's datetime or sortable string)
    all_data.sort(key=lambda x: x[0] if isinstance(x[0], datetime) else datetime.min)

    # Write to JULIJ 26 sheet
    if 'JULIJ 26' in wb.sheetnames:
        ws_julij = wb['JULIJ 26']
        # Clear existing data
        ws_julij.delete_rows(1, ws_julij.max_row)
    else:
        ws_julij = wb.create_sheet('JULIJ 26')
        
    headers = ["Datum", "Tip", "Šifra materiala", "Naziv", "Količina", "Dodatni podatki", "Opomba"]
    ws_julij.append(headers)
    
    # Format headers
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws_julij[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Append data
    for row_data in all_data:
        # If datum is datetime, format it to string for nice display, or just keep as datetime (excel will format)
        if isinstance(row_data[0], datetime):
            row_data[0] = row_data[0].strftime("%d. %m. %Y")
        ws_julij.append(row_data)
        
    # Adjust column widths
    ws_julij.column_dimensions['A'].width = 15
    ws_julij.column_dimensions['B'].width = 12
    ws_julij.column_dimensions['C'].width = 15
    ws_julij.column_dimensions['D'].width = 40
    ws_julij.column_dimensions['E'].width = 12
    ws_julij.column_dimensions['F'].width = 35
    ws_julij.column_dimensions['G'].width = 25

    # Save
    wb.save(file_path)
    print(f"Uspešno prepisano {len(all_data)} vrstic za julij v list 'JULIJ 26'.")

if __name__ == '__main__':
    main()
