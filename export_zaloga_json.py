import openpyxl
import json
import os
from datetime import datetime

def export_zaloga():
    xlsx_path = r'\\server-2012\SIMON\ZALOGA MATERIALA 2026.xlsx'
    if not os.path.exists(xlsx_path):
        current_year = datetime.now().year
        xlsx_path = r'\\server-2012\SIMON\ZALOGA MATERIALA {}.xlsx'.format(current_year)
    
    if not os.path.exists(xlsx_path):
        print(f"Napaka: Datoteka {xlsx_path} ne obstaja.")
        return

    print(f"Berem datoteko: {xlsx_path} ...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Zberemo prispelo po šifrah
    prispelo_map = {}
    if 'PRISPELO' in wb.sheetnames:
        ws_prispelo = wb['PRISPELO']
        for r in range(2, ws_prispelo.max_row + 1):
            sifra = ws_prispelo.cell(r, 2).value
            kol = ws_prispelo.cell(r, 7).value
            if sifra is not None and kol is not None:
                try:
                    s_str = str(int(sifra) if isinstance(sifra, (int, float)) else sifra).strip()
                    prispelo_map[s_str] = prispelo_map.get(s_str, 0.0) + float(kol)
                except (ValueError, TypeError):
                    pass

    # Zberemo odpis po šifrah
    odpis_map = {}
    if 'ODPIS' in wb.sheetnames:
        ws_odpis = wb['ODPIS']
        for r in range(2, ws_odpis.max_row + 1):
            sifra = ws_odpis.cell(r, 2).value
            kol = ws_odpis.cell(r, 7).value
            if sifra is not None and kol is not None:
                try:
                    s_str = str(int(sifra) if isinstance(sifra, (int, float)) else sifra).strip()
                    odpis_map[s_str] = odpis_map.get(s_str, 0.0) + float(kol)
                except (ValueError, TypeError):
                    pass

    ws_zaloga = wb['TRENUTNA ZALOGA']
    stock_data = []

    for r in range(2, ws_zaloga.max_row + 1):
        sifra = ws_zaloga.cell(r, 1).value
        if sifra is None:
            continue
        s_str = str(int(sifra) if isinstance(sifra, (int, float)) else sifra).strip()
        naziv = (ws_zaloga.cell(r, 2).value or '').strip()
        sirina = ws_zaloga.cell(r, 3).value or ''
        visina = ws_zaloga.cell(r, 4).value or ''
        gramatura = ws_zaloga.cell(r, 5).value or ''
        zacetna = ws_zaloga.cell(r, 6).value or 0
        try:
            zacetna_val = float(zacetna)
        except (ValueError, TypeError):
            zacetna_val = 0.0
            
        p_val = prispelo_map.get(s_str, 0.0)
        o_val = odpis_map.get(s_str, 0.0)
        zaloga_val = zacetna_val + p_val - o_val
        
        format_str = f"{sirina}x{visina}" if sirina and visina else ""
        
        item = {
            'šifra materiala': s_str,
            'sifra': s_str,
            'naziv': naziv,
            'širina': sirina,
            'višina': visina,
            'format': format_str,
            'gramatura': gramatura,
            'začetna zaloga': zacetna_val,
            'prispelo': p_val,
            'odpis': o_val,
            'zaloga': zaloga_val
        }
        stock_data.append(item)

    out_local = r'c:\DARKO\KalkulacijaPetric\zalogaSimon.json'
    out_server = r'\\server-2012\SIMON\zalogaSimon.json'

    with open(out_local, 'w', encoding='utf-8') as f:
        json.dump(stock_data, f, ensure_ascii=False, indent=2)
    print(f"Izvoženo {len(stock_data)} artiklov v {out_local}")

    try:
        with open(out_server, 'w', encoding='utf-8') as f:
            json.dump(stock_data, f, ensure_ascii=False, indent=2)
        print(f"Izvoženo {len(stock_data)} artiklov v {out_server}")
    except Exception as e:
        print(f"Opozorilo pri zapisovanju na strežnik: {e}")

if __name__ == '__main__':
    export_zaloga()
