import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Load existing data
wb_old = openpyxl.load_workbook("d:\\Git\\KalkulacijaPetric\\zalogaSimon.xlsx", data_only=True)
ws_old = wb_old['test']

# Create new workbook
wb_new = openpyxl.Workbook()
ws_zaloga = wb_new.active
ws_zaloga.title = "TRENUTNA ZALOGA"
ws_prispelo = wb_new.create_sheet("PRISPELO")
ws_odpis = wb_new.create_sheet("ODPIS")

# --- FORMATTING ---
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

def style_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

# --- SHEET 1: ZALOGA ---
headers_zaloga = [
    "Šifra materiala", "Naziv materiala", "Širina", "Višina", "Gramatura", 
    "Začetna zaloga", "Skupaj Prispelo", "Skupaj Odpisano", "TRENUTNA ZALOGA"
]
style_header(ws_zaloga, headers_zaloga)

row_idx = 2
for row in ws_old.iter_rows(min_row=2, values_only=True):
    if not row[0]: # Skip empty rows
        continue
    # Extract existing data
    sifra = row[0]
    ime = row[1]
    sirina = row[2]
    visina = row[3]
    gram = row[4]
    zacetna = row[5] if row[5] is not None else 0
    
    # Write to new sheet
    ws_zaloga.cell(row=row_idx, column=1, value=sifra)
    ws_zaloga.cell(row=row_idx, column=2, value=ime)
    ws_zaloga.cell(row=row_idx, column=3, value=sirina)
    ws_zaloga.cell(row=row_idx, column=4, value=visina)
    ws_zaloga.cell(row=row_idx, column=5, value=gram)
    ws_zaloga.cell(row=row_idx, column=6, value=zacetna)
    
    # Formulas
    # Skupaj Prispelo = SUMIFS(PRISPELO!D:D, PRISPELO!B:B, A2) -> sum Količina (D) where Šifra (B) matches
    ws_zaloga.cell(row=row_idx, column=7, value=f'=SUMIF(PRISPELO!B:B, A{row_idx}, PRISPELO!D:D)')
    # Skupaj Odpisano = SUMIFS(ODPIS!D:D, ODPIS!B:B, A2) -> sum Količina (D) where Šifra (B) matches
    ws_zaloga.cell(row=row_idx, column=8, value=f'=SUMIF(ODPIS!B:B, A{row_idx}, ODPIS!D:D)')
    # Trenutna zaloga = F2 + G2 - H2
    cell_trenutna = ws_zaloga.cell(row=row_idx, column=9, value=f'=F{row_idx} + G{row_idx} - H{row_idx}')
    cell_trenutna.font = Font(bold=True)
    
    row_idx += 1

# Adjust column widths
for col in ['A', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
    ws_zaloga.column_dimensions[col].width = 15
ws_zaloga.column_dimensions['B'].width = 35

# --- SHEET 2: PRISPELO ---
headers_prispelo = ["Datum", "Šifra materiala", "Naziv (Avtomatsko)", "Količina (Prejem)", "Št. Dobavnice", "Dobavitelj", "Opomba"]
style_header(ws_prispelo, headers_prispelo)

for i in range(2, 1000):
    # VLOOKUP for Name: =IF(B2="","",VLOOKUP(B2,'TRENUTNA ZALOGA'!A:B,2,FALSE))
    ws_prispelo.cell(row=i, column=3, value=f'=IF(B{i}="","",VLOOKUP(B{i},\'TRENUTNA ZALOGA\'!A:B,2,FALSE))')

for col in ['A', 'B', 'D', 'E', 'F', 'G']:
    ws_prispelo.column_dimensions[col].width = 15
ws_prispelo.column_dimensions['C'].width = 35

# --- SHEET 3: ODPIS ---
headers_odpis = ["Datum", "Šifra materiala", "Naziv (Avtomatsko)", "Količina (Odpis)", "Delovni Nalog", "Oseba", "Opomba"]
style_header(ws_odpis, headers_odpis)

for i in range(2, 1000):
    ws_odpis.cell(row=i, column=3, value=f'=IF(B{i}="","",VLOOKUP(B{i},\'TRENUTNA ZALOGA\'!A:B,2,FALSE))')

for col in ['A', 'B', 'D', 'E', 'F', 'G']:
    ws_odpis.column_dimensions[col].width = 15
ws_odpis.column_dimensions['C'].width = 35

# Save
new_path = "d:\\Git\\KalkulacijaPetric\\Zaloga_Avtomatizirana.xlsx"
wb_new.save(new_path)
print(f"Created successfully at {new_path}")
