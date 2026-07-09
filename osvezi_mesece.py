import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
import sys

MONTHS_SLO = {
    1: "JANUAR", 2: "FEBRUAR", 3: "MAREC", 4: "APRIL",
    5: "MAJ", 6: "JUNIJ", 7: "JULIJ", 8: "AVGUST",
    9: "SEPTEMBER", 10: "OKTOBER", 11: "NOVEMBER", 12: "DECEMBER"
}

def parse_date(datum_val):
    if isinstance(datum_val, datetime):
        return datum_val
    if isinstance(datum_val, str):
        # poskusimo razčleniti različne formate
        datum_val = datum_val.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y", "%d. %m. %Y", "%d.%m.%y"):
            try:
                return datetime.strptime(datum_val, fmt)
            except ValueError:
                pass
    return None

def main():
    file_path = r"S:\SIMON\Zaloga_Avtomatizirana.xlsx"
    print("Berem datoteko s serverja S (S:\SIMON)...")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
    except PermissionError:
        print("NAPAKA: Datoteka je odprta v Excelu! Prosim zaprite jo in poskusite znova.")
        input("Pritisnite Enter za izhod...")
        sys.exit(1)
        
    ws_zaloga = wb_data['TRENUTNA ZALOGA']
    naziv_dict = {}
    for row in ws_zaloga.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            naziv_dict[str(row[0])] = row[1]
            
    ws_prispelo = wb_data['PRISPELO']
    ws_odpis = wb_data['ODPIS']
    
    monthly_data = {} # { "JULIJ 26": [rows...], "AVGUST 26": [rows...] }
    
    def process_sheet(ws, tip, extra_cols):
        for row in ws.iter_rows(min_row=2, values_only=True):
            datum_raw = row[0]
            if datum_raw is None or str(datum_raw).strip() == "":
                continue
                
            dt = parse_date(datum_raw)
            if dt is None:
                continue
                
            month_name = MONTHS_SLO.get(dt.month, "NEZNANO")
            year_short = str(dt.year)[-2:]
            sheet_name = f"{month_name} {year_short}"
            
            if sheet_name not in monthly_data:
                monthly_data[sheet_name] = []
                
            sifra = str(row[1]) if row[1] is not None else ""
            naziv = naziv_dict.get(sifra, row[2] or "")
            kolicina = row[3]
            
            # tip=PRISPELO -> 4:Št. Dobavnice, 5:Dobavitelj
            # tip=ODPIS -> 4:Delovni Nalog, 5:Oseba
            dodatno = ""
            if tip == "PRISPELO":
                dodatno = f"Dobavnica: {row[4] or ''}, Dobavitelj: {row[5] or ''}"
            else:
                dodatno = f"DN: {row[4] or ''}, Oseba: {row[5] or ''}"
                
            opomba = row[6] or ""
            
            monthly_data[sheet_name].append([dt, tip, sifra, naziv, kolicina, dodatno, opomba])

    process_sheet(ws_prispelo, "PRISPELO", 4)
    process_sheet(ws_odpis, "ODPIS", 4)
    
    if not monthly_data:
        print("Ni bilo najdenih ustreznih datumov za razvrščanje.")
        input("Pritisnite Enter za izhod...")
        return
        
    for sheet_name, rows in monthly_data.items():
        # sortiramo po datumu
        rows.sort(key=lambda x: x[0])
        
        if sheet_name in wb.sheetnames:
            ws_month = wb[sheet_name]
            ws_month.delete_rows(1, ws_month.max_row)
        else:
            ws_month = wb.create_sheet(sheet_name)
            
        headers = ["Datum", "Tip", "Šifra materiala", "Naziv", "Količina", "Dodatni podatki", "Opomba"]
        ws_month.append(headers)
        
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws_month[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for r in rows:
            # format date beautifully
            formatted_date = r[0].strftime("%d. %m. %Y")
            row_to_write = [formatted_date] + r[1:]
            ws_month.append(row_to_write)
            
        # Adjust column widths
        ws_month.column_dimensions['A'].width = 15
        ws_month.column_dimensions['B'].width = 12
        ws_month.column_dimensions['C'].width = 15
        ws_month.column_dimensions['D'].width = 40
        ws_month.column_dimensions['E'].width = 12
        ws_month.column_dimensions['F'].width = 35
        ws_month.column_dimensions['G'].width = 25
        
        print(f"Osvežen zavihek: {sheet_name} ({len(rows)} knjižb)")

    print("Shranjujem datoteko...")
    wb.save(file_path)
    print("USPEŠNO ZAKLJUČENO!")
    
if __name__ == '__main__':
    main()
