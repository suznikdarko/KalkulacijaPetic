import sys

try:
    import pandas as pd
    file_path = "d:\\Git\\KalkulacijaPetric\\zalogaSimon.xlsx"
    xl = pd.ExcelFile(file_path)
    print("Sheets:", xl.sheet_names)
    
    for sheet in xl.sheet_names:
        df = xl.parse(sheet, nrows=5)
        print(f"\n--- Sheet: {sheet} ---")
        print(df.columns.tolist())
        print(df.head())
except Exception as e:
    print("Error with pandas:", e)
    # Fallback to openpyxl if pandas fails
    try:
        import openpyxl
        file_path = "d:\\Git\\KalkulacijaPetric\\zalogaSimon.xlsx"
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print("\nFallback openpyxl sheets:", wb.sheetnames)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            print(f"--- Sheet: {sheet} ---")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                print(row)
                if i >= 5:
                    break
    except Exception as e2:
        print("Error with openpyxl:", e2)
