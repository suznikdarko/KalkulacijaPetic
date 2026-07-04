import sys
import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Font

try:
    file_path = "d:\\Git\\KalkulacijaPetric\\Zaloga_Avtomatizirana.xlsx"
    wb = openpyxl.load_workbook(file_path)
    ws = wb['TRENUTNA ZALOGA']

    # Ustvarimo rdeče ozadje in bel, krepak font
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True)

    # Pravilo: Če celica A (šifra) ni prazna IN je zaloga (I) manjša od 100
    rule = FormulaRule(formula=['AND(A2<>"", I2<100)'], stopIfTrue=True, fill=red_fill, font=white_font)

    # Dodamo pogojno oblikovanje za stolpec I (od vrstice 2 do 1000)
    ws.conditional_formatting.add('I2:I1000', rule)

    wb.save(file_path)
    print("Successfully added conditional formatting!")
except Exception as e:
    print("Error:", e)
