import sys
import openpyxl

try:
    file_path = "d:\\Git\\KalkulacijaPetric\\Zaloga_Avtomatizirana.xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws_odpis = wb['ODPIS']
    
    print("--- ODPIS SHEET ROWS ---")
    rows = list(ws_odpis.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        print(row)
        
    print("\n--- TRENUTNA ZALOGA (First 5 rows) ---")
    ws_zaloga = wb['TRENUTNA ZALOGA']
    for i, row in enumerate(ws_zaloga.iter_rows(values_only=True)):
        print(row)
        if i >= 5:
            break
except Exception as e:
    print("Error:", e)
