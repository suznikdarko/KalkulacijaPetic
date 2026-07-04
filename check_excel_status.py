import sys
import openpyxl

try:
    file_path = "d:\\Git\\KalkulacijaPetric\\Zaloga_Avtomatizirana.xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws_odpis = wb['ODPIS']
    
    print("--- ODPIS SHEET ROWS ---")
    rows = list(ws_odpis.iter_rows(values_only=True))
    for i, row in enumerate(rows):
        if any(row):  # only print non-empty rows
            print(row)
        
    print("\n--- TRENUTNA ZALOGA (First 10 rows) ---")
    ws_zaloga = wb['TRENUTNA ZALOGA']
    for i, row in enumerate(ws_zaloga.iter_rows(values_only=True)):
        if row[7] != 0 and row[7] is not None and row[7] != 'Skupaj Odpisano':  # check 'Skupaj Odpisano' column
            print(f"Row {i+1}: {row}")
        elif i < 5:
            print(f"Row {i+1}: {row}")
            
except Exception as e:
    print("Error:", e)
